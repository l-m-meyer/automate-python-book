import openpyxl

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']            # Get a sheet from the workbook
print(sheet['A1'])              # Get a cell from the sheet
# <Cell 'Sheet1'.A1>

print(sheet['A1'].value)        # Get the value from the cell
# datetime.datetime(2015, 4, 5, 13, 34, 2)

c = sheet['B1']                 # Get another cell from the sheet
print(c.value)
# 'Apples'

# Get the row, column, and value from the cell
print(f'Row {c.row}, Column {c.column} is {c.value}')
# 'Row 1, Column B is Apples'

print(f'Cell {c.coordinate} is {c.value}')
# 'Cell B1 is Apples'

print(sheet['C1'].value)
# 73

print(sheet.cell(row=1, column=2))
# <Cell 'Sheet1'.B1>

print(sheet.cell(row=1, column=2))
# 'Apples'

for i in range(1, 8, 2):        # Go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries


print(sheet.max_row)            # Get the highest row number
# 7

print(sheet.max_column)         # Get the highest column number
# 3