import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


get_column_letter(1)                # Translate column 1 to a letter.
# 'A'

get_column_letter(2)
# 'B'

get_column_letter(27)
# 'AA'

get_column_letter(900)
# 'AHP'

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
# 'C'

column_index_from_string('A')       # Get A's number
# 1

column_index_from_string('AA')
# 27