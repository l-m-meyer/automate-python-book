import openpyxl


# Create a new workbook
wb = openpyxl.Workbook()
print(wb.sheetnames)


# Add a new sheet
wb.create_sheet()
print(wb.sheetnames)


# Create new sheet at index 0
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)


# Create a new sheet at index 2
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)


# Delete sheets from workbook
del wb['Middle Sheet']
del wb['Sheet1']
print(wb.sheetnames)


# Save workbook
# wb.save('sheet_exercise.xlsx')    # uncomment to save