import openpyxl


# CREATE NEW WORKBOOK
wb = openpyxl.Workbook()                # Create a blank workbook.
print(wb.sheetnames)                    # Starts with one sheet: ['Sheet']

sheet = wb.active
print(sheet.title)                      # 'Sheet'

sheet.title = 'Spam Bacon Eggs Sheet'   # Change title
print(wb.sheetnames)                    # ['Spam Bacon Eggs Sheet']


# EDIT EXISTING WORKBOOK
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam Spam'
wb.save('example_copy.xlsx')            # Save the workbook