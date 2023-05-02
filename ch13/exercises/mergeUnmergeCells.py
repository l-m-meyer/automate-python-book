import openpyxl


# create a new workbook
wb = openpyxl.Workbook()
sheet = wb.active


# merge cells
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together.'

sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells.'


# save workbook
wb.save('merged.xlsx')


# unmerge cells
sheet.unmerge_cells('A1:D3')    # Split up cells
sheet.unmerge_cells('C5:D5')


# save workbook
wb.save('merged.xlsx')