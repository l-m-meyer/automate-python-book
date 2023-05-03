#! python3
# invertCells.py - Inverts the row and column of the cells in a spreadsheet.


import openpyxl, random
from openpyxl.utils import get_column_letter


def create_sample_data():
    # create blank workbook
    wb = openpyxl.Workbook()
    sheet = wb.active


    # enter data into blank workbook
    sheet['A1'] = 'ITEM'
    sheet['B1'] = 'SOLD'

    items = ['Eggplant', 'Cucumber', 'Cabbage', 'Garlic', 'Parsnips', 'Asparagus', 'Brussel Sprouts']

    for i, item in enumerate(items):
        ran = random.randint(1, 500)
        
        sheet[f'A{i + 2}'].value = item
        sheet[f'B{i + 2}'].value = ran


    # save workbook
    filename = 'sampledata.xlsx'
    wb.save(filename)

    return filename


# invert cells from sheet[x][y] to sheet[y][x]
def invert_cells(sheet, copy):
    for r in tuple(sheet.rows):
        for c in r:
            col = get_column_letter(c.row)
            row = c.column
            
            copy[f'{col}{row}'].value = c.value
            

# load file and copy it to invert
def copy_file(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    wb_copy = openpyxl.Workbook()
    sheet_copy = wb_copy.active

    return sheet, sheet_copy, wb_copy


def main():
    filename = create_sample_data()
    sheet, sheet_copy, wb_copy = copy_file(filename)
    invert_cells(sheet, sheet_copy)
    wb_copy.save('inverteddata.xlsx')


if __name__ == '__main__':
    main()