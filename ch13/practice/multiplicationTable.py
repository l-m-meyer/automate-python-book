#! python3
# multiplicationTable.py - Creates a NxN multiplication table in an Excel spreadsheet.


import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pyinputplus as pyip


# create blank workbook
wb = openpyxl.Workbook()
sheet = wb.active


# get number N from the command line
try:
    n = sys.argv[1]
except IndexError as e:
    print('Forget something? A number perhaps?\nError:', e)
    print()
    print('Try again...')
    n = pyip.inputInt('Enter a number: ')


# create font object headers
headerFont = Font(size=12, bold=True)


# create column and row headers in bold
for i in range(2, int(n) + 2):
    # columns
    sheet.cell(row=1, column=i).value = i - 1
    sheet.cell(row=1, column=i).font = headerFont

    # rows
    sheet.cell(row=i, column=1).value = i - 1
    sheet.cell(row=i, column=1).font = headerFont


# add formula to calculate values for each cell
for i in range(2, int(n) + 2):
    col = get_column_letter(i)
    for j in range(2, int(n) + 2):
        formula = f'={col}1 * A{j}'
        sheet[f'{col}{j}'].value = formula


# save workbook
wb.save('multiplicationTable.xlsx')