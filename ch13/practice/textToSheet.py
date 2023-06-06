#! python3
# textToSheet.py - Reads the contents of several text files and inserts those contents into a spreadsheet.


import os
import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


def create_workbook():
    # create new workbook
    wb = openpyxl.Workbook()
    sheet = wb.active

    return wb, sheet


def check_if_exists(files):
    for file in files:
        # removes nonexistent files from files list
        if not os.path.exists(file):
            files.remove(file)
    
    return files


def get_files():
    # get files from command line
    args = sys.argv[1:]
    
    # checks that there is at least one command line argument in args list
    if not args:
        args = input("Please include at least one '.txt' file:\n").split(' ')

    # filters list for only '.txt' files
    files = list(filter(lambda f: f.endswith('.txt'), args))

    # checks that there is at least one '.txt' file in files list
    if not files:
        get_files()

    # checks that files exist
    files = check_if_exists(files)
   
    return files


def get_current_column(column_letter):
    # if column_letter has not been defined, set to column A
    if not column_letter:
        return get_column_letter(1)
    
    # adds 1 to current column index
    column_index = column_index_from_string(column_letter) + 1

    return get_column_letter(column_index)

    
def process_files(files, sheet):
    column_letter = ''

    for file in files:
        # open each file
        with open(file, encoding='utf-8') as f:
            # read each line of file to get a list of strings
            lines = f.read().splitlines()

            # get current column letter
            column_letter = get_current_column(column_letter)
            
            # writes each string in list to new row of same column
            for i, line in enumerate(lines, start=1):
                sheet[f'{column_letter}{i}'].value = line


def main():
    wb, sheet = create_workbook()
    files = get_files()
    process_files(files, sheet)
    wb.save('textToSheet.xlsx')


if __name__ == '__main__':
    main()